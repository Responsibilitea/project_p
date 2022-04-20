# 利用python将excel文件Sales Summary demo.xlsm中所有数字大于9000的cell highlight为红色背景
# 完成后重新存为Sales Summary demo1.xlsx

from openpyxl import load_workbook
from openpyxl.styles import PatternFill

workbook_1 = load_workbook("E:\\file\\excel\\test\\Sales Summary demo.xlsm")
red_fill = PatternFill(fill_type="solid", start_color="FFFF0000", end_color="FFFF0000")
sheet_list = []

for ws in workbook_1:
    sheet_list.append(ws)

for ws in sheet_list:
    for row in ws:
        for cell in row:
            if isinstance(cell.value, int) and cell.value > 9000:
                print(cell.value)
                cell.fill = red_fill

workbook_1.save("E:\\file\\excel\\test\\Sales Summary demo1.xlsx")

workbook_1.close()
